# IMPORTS!
import datetime
import random
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db import connection
# MODELS
from .models import Estudiante, Inscription, Pago
from .forms import Create_EstudianteForm, Create_PagoForm, Create_InscriptionForm
from .forms import Update_EstudianteForm, Update_PagoForm, Update_InscriptionForm

# TOOLS
from django.core.paginator import Paginator
from openpyxl import Workbook
from num2words import num2words
import re
# Create your views here.

from pathlib import Path
@login_required(login_url='signin')
def main(request):
    BASE_DIR = Path(__file__).resolve().parent.parent
    print(BASE_DIR)
    dictionary = {
        'tittle': 'Main',
        'return_url': 'main',
    }
    return render(request, 'inicio.html', dictionary)

# SECCION DE CREACION


@login_required(login_url='signin')
def create_student(request):
    if request.method == 'GET':
        dictionary = {
            'tittle': 'Creacion de Estudiante',
            'form': Create_EstudianteForm(),
            'return_url': 'main',
        }
        return render(request, 'form_create.html', dictionary)
    else:
        catch_form = Create_EstudianteForm(request.POST)
        save_form = catch_form.save(commit=False)
        save_form.save()
        # url = f'create_inscription/{save_form.dni}'
        return redirect('create_inscription', save_form.dni)


@login_required(login_url='signin')
def create_inscription(request, student_id):
    if request.method == 'GET':
        dictionary = {
            'tittle': 'Creacion de Inscripcion',
            'form': Create_InscriptionForm,
            'return_url': 'main',
        }
        return render(request, 'form_create.html', dictionary)
    else:
        consulta = Estudiante.objects.get(dni=student_id)
        form = Create_InscriptionForm(request.POST)
        post_form = form.save(commit=False)
        post_form.fk_estudiante_id = consulta.pk
        post_form.num_comprobante = f'FMS-{random.randint(100, 999)}-{random.randint(100, 999)}'
        post_form.save()
        return redirect('create_pay', post_form.num_comprobante)


@login_required(login_url='signin')
def create_pay(request, enrollment_id):
    if request.method == 'GET':
        dictionary = {
            'tittle': 'Creacion de pago',
            'form': Create_PagoForm,
            'return_url': 'main',
        }
        return render(request, 'form_create.html', dictionary)
    else:
        form = Create_PagoForm(request.POST)
        post_form = form.save(commit=False)
        post_form.fk_inscription_id = enrollment_id
        post_form.factura_pago = f'FPS-{random.randint(100, 999)}-{random.randint(100, 999)}'
        post_form.save()
        return redirect('list_pays', post_form.fk_inscription_id)


def add_pass(request, num_pass_id):
    obj = get_object_or_404(Pago, pk=num_pass_id)
    obj.monto_valor = obj.monto_valor-obj.monto_pagar
    obj.monto_pagar = 0.00
    obj.fecha_pago = datetime.datetime.now()
    sufijo_re = re.compile(r"_\d{2}$")
    # Buscar un sufijo en el número de factura
    match = sufijo_re.search(f"{obj.factura_pago}")
    if match:
        # Si hay un sufijo, incrementar en 1 el número siguiente al guion bajo
        num_sufijo = int(match.group()[1:]) + 1
        nuevo_num_fact = obj.factura_pago[:-2] + f"{num_sufijo:02d}"
    else:
        # De lo contrario, agregar "_01" al final del número de factura
        nuevo_num_fact = obj.factura_pago + "_01"
    # Imprimir el nuevo número de factura
    obj.factura_pago = nuevo_num_fact
    obj_form = Create_PagoForm(instance=obj)
    if request.method == 'GET':
        dictionary = {
            'tittle': 'Agregar Abono',
            'form': obj_form,
            # DETALLES DE URLS
            'has_id': True,
            'return_url': 'opciones_pago',
            'return_id': num_pass_id,
        }
        return render(request, 'form_create.html', dictionary)
    else:
        obj_form2 = Create_PagoForm(request.POST, instance=obj)
        if obj_form2.is_valid():
            obj_form2.save()
            return redirect('list_pays', obj.fk_inscription.num_comprobante)
        else:
            return HttpResponse('error desconocido')

# SECCION DE LISTADOS


@login_required(login_url='signin')
def list_students(request):
    # FILTER_SQL
    if request.method == 'GET':
        consulta = f"SELECT dni, apellido_1, nombre_1, sexo, phone from app_estudiante;"
    else:
        consulta = f"SELECT dni, apellido_1, nombre_1, sexo, phone from app_estudiante where dni like '%{request.POST.get('search_filter')}%';"
    # END_FILTER_SQL

    # SQL_CONNECTION
    sql_connection = connection.cursor()
    sql_connection.execute(consulta)
    result = sql_connection.fetchall()
    sql_connection.close()
    # END_SQL_CONNECTION
    # BLOCK_PAGINACION
    paginator = Paginator(result, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    # END_PAGINACION

    if len(result) > 0:
        dictionary = {
            'tittle': 'Listado de Estudiantes',
            'labels': ('Cedula', 'Apellido', 'Nombre', 'Sexo', 'Telefono'),
            'items_data': page_obj,
            'return_url': 'main',
            'redirect_url': 'opciones_estudiante',
            'query': request.POST.get('search_filter'),
        }
        return render(request, 'list_all.html', dictionary)
    else:
        dictionary = {
            'tittle': 'Creacion de Estudiante',
            'error': 'No existen estudiante registrados',
        }
        return redirect('create_students')


@login_required(login_url='signin')
def list_inscriptions(request, student_id):
    obj = get_object_or_404(Estudiante, pk=student_id)
    # FILTER_SQL
    if request.method == 'GET':
        consulta = f"select i.num_comprobante, i.date_year, i.curso, i.curso_nivel, i.trimestre from app_inscription as i inner join app_estudiante as e on i.fk_estudiante_id = e.dni and e.dni = '{student_id}' order by i.date_year, i.curso, i.curso_nivel, i.trimestre desc;"
    else:
        consulta = f"select i.num_comprobante, i.date_year, i.curso, i.curso_nivel, i.trimestre from app_inscription as i inner join app_estudiante as e on i.fk_estudiante_id = e.dni and e.dni = '{student_id}' where i.num_comprobante like '%{request.POST.get('search_filter')}%';"
    # END_FILTER_SQL

    # SQL_CONNECTION
    sql_connection = connection.cursor()
    sql_connection.execute(consulta)
    result = sql_connection.fetchall()
    sql_connection.close()
    # END_SQL_CONNECTION
    # BLOCK_PAGINACION
    paginator = Paginator(result, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    # END_PAGINACION

    if len(result) > 0:
        dictionary = {
            'estudiante': f'{obj.apellido_1}, {obj.nombre_1}',
            'estudiante_dni': f'{obj.dni}',
            'tittle': 'Listado de Inscripciones',
            'labels': ('comprobante', 'año', 'curso', 'nivel del curso', 'trimestre'),
            'items_data': page_obj,
            'query': request.POST.get('search_filter'),

            # DETALLES DE URLS
            'has_id': True,
            'return_url': 'opciones_estudiante',
            'return_id': student_id,
            'redirect_url': 'opciones_inscripcion',
        }
        return render(request, 'list_all.html', dictionary)
    else:
        return redirect('create_inscription', student_id)


@login_required(login_url='signin')
def list_pays(request, enrollment_id):
    obj = get_object_or_404(Inscription, pk=enrollment_id)
    # FILTER_orm

    # FILTER_SQL
    if request.method == 'GET':
        consulta = f"select factura_pago, cargo_tipo, ROUND(monto_valor, 2), ROUND(monto_pagar, 2), case when monto_valor-monto_pagar > 0 then -(round(monto_valor-monto_pagar,2)) else round(monto_valor-monto_pagar,2) end as SALDO, fecha_pago from app_pago inner join app_inscription on fk_inscription_id = num_comprobante and num_comprobante = '{enrollment_id}' order by cargo_tipo desc, factura_pago desc;"

    else:
        consulta = f"select factura_pago, cargo_tipo, monto_valor, monto_pagar, case when monto_valor-monto_pagar > 0 then -(round(monto_valor-monto_pagar,2)) else round(monto_valor-monto_pagar,2) end as saldo, fecha_pago from app_pago inner join app_inscription on fk_inscription_id = num_comprobante and num_comprobante = '{enrollment_id}' where factura_pago like '%{request.POST.get('search_filter')}%' order by cargo_tipo desc, factura_pago desc;"
    # END FILTER_SQL

    # SQL_CONNECTION
    sql_connection = connection.cursor()
    sql_connection.execute(consulta)
    result = sql_connection.fetchall()
    sql_connection.close()
    # END_SQL_CONNECTION
    # BLOCK_PAGINACION
    paginator = Paginator(result, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    # END_PAGINACION

    if len(result) > 0:
        dictionary = {
            'estudiante': f'{obj.fk_estudiante.apellido_1}, {obj.fk_estudiante.nombre_1}',
            'estudiante_dni': f'{obj.fk_estudiante.dni}',
            'tittle': 'Listado de Pagos',
            'labels': ('Pago', 'tipo de cargo', 'Monto', 'Pagó', 'Saldo', 'Fecha de Pago', ),
            'items_data': page_obj,
            'redirect_url': 'opciones_pago',
            'query': request.POST.get('search_filter'),
            # DETALLES DE URLS
            'has_id': True,
            'return_url': 'opciones_inscripcion',
            'return_id': obj.num_comprobante,
        }
        return render(request, 'list_all.html', dictionary)
    else:
        return redirect('create_pay', enrollment_id)

# SECCION DE OPCIONES


@login_required(login_url='signin')
def option_student(request, student_id):
    return render(request, 'options/student.html', {
        'id': student_id,
        'return_url': 'list_students',
    })


@login_required(login_url='signin')
def option_inscription(request, enrollment_id):
    obj = get_object_or_404(Inscription, pk=enrollment_id)
    return render(request, 'options/inscription.html', {
        'id': enrollment_id,
        # DETALLES DE URLS
        'has_id': True,
        'return_url': 'list_inscriptions',
        'return_id': obj.fk_estudiante.dni,
    })


@login_required(login_url='signin')
def option_pays(request, num_fact):
    obj = get_object_or_404(Pago, pk=num_fact)
    return render(request, 'options/pays.html', {
        'id': num_fact,
        # DETALLES DE URLS
        'has_id': True,
        'return_url': 'list_pays',
        'return_id': obj.fk_inscription.num_comprobante,
    })

# SECCION DE ACTUALIZACION


@login_required(login_url='signin')
def update_student(request, student_id):
    obj = get_object_or_404(Estudiante, pk=student_id)
    obj_form = Update_EstudianteForm(instance=obj)
    if request.method == 'GET':
        dictionary = {
            'tittle': 'Actualizar Estudiante',
            'form': obj_form,
            # DETALLES DE URLS
            'has_id': True,
            'return_url': 'opciones_estudiante',
            'return_id': student_id,
        }
        return render(request, 'form_create.html', dictionary)
    else:
        obj_form2 = Update_EstudianteForm(request.POST, instance=obj)
        if obj_form2.is_valid():
            obj_form2.save()
            return redirect('list_students')
        else:
            return redirect('create_student')


@login_required(login_url='signin')
def update_inscription(request, enrollment_id):
    obj = get_object_or_404(Inscription, pk=enrollment_id)
    obj_form = Update_InscriptionForm(instance=obj)
    if request.method == 'GET':
        dictionary = {
            'tittle': 'Actualizar Inscripcion',
            'form': obj_form,
            # DETALLES DE URLS
            'has_id': True,
            'return_url': 'opciones_inscripcion',
            'return_id': enrollment_id,
        }
        return render(request, 'form_create.html', dictionary)
    else:
        obj_form2 = Update_InscriptionForm(request.POST, instance=obj)
        if obj_form2.is_valid():
            obj_form2.save()
            return redirect('list_inscriptions', obj.fk_estudiante.dni)
        else:
            return redirect('create_inscription', obj.fk_estudiante.dni)


@login_required(login_url='signin')
def update_pay(request, num_fact_id):
    obj = get_object_or_404(Pago, pk=num_fact_id)
    obj_form = Update_PagoForm(instance=obj)
    if request.method == 'GET':
        dictionary = {
            'tittle': 'Actualizar Pago',
            'form': obj_form,
            # DETALLES DE URLS
            'has_id': True,
            'return_url': 'opciones_pago',
            'return_id': num_fact_id,
        }
        return render(request, 'form_create.html', dictionary)
    else:
        obj_form2 = Update_PagoForm(request.POST, instance=obj)
        if obj_form2.is_valid():
            obj_form2.save()
            return redirect('list_pays', obj.fk_inscription.num_comprobante)
        else:
            return redirect('create_pay', obj.fk_inscription.num_comprobante)

# SECCION DE ELIMINACION


@login_required(login_url='signin')
def delete_student(request, student_id):
    obj = get_object_or_404(Estudiante, pk=student_id)
    obj.delete()
    return redirect('list_students')


@login_required(login_url='signin')
def delete_inscription(request, enrollment_id):
    obj = get_object_or_404(Inscription, pk=enrollment_id)
    obj.delete()
    return redirect('list_inscriptions', obj.fk_estudiante.dni)


@login_required(login_url='signin')
def delete_pay(request, num_fact_id):
    obj = get_object_or_404(Pago, pk=num_fact_id)
    obj.delete()
    return redirect('list_pays', obj.fk_inscription.num_comprobante)

# IMPRESION DE PAGOS POR MATRICULA


@login_required(login_url='signin')
def print_fact(request, num_fact_id):
    obj = get_object_or_404(Pago, pk=num_fact_id)
    wb = Workbook()
    ws = wb.active

    # Escribimos los títulos de las columnas
    ws['B2'] = 'INSTITUTO GIORDANO BRUNO'
    ws.merge_cells('B2:F2')
    ws['B3'] = 'R.U.C. 443964-1-430566 D.V. 65'
    ws.merge_cells('B3:F3')
    ws['B4'] = 'E-Mail: institutogiordanobruno2004@gmail.com'
    ws.merge_cells('B4:F4')
    # ----------------------------------------------------------------
    ws['B7'] = 'RECIBO N°:'
    ws['C7'] = f'{obj.factura_pago}'
    ws['E7'] = 'Fecha:'
    ws['F7'] = f'{obj.fecha_pago}'
    # ----------------------------------------------------------------
    ws['B9'] = 'Hemos Recibido de:'
    ws['D9'] = f'{obj.fk_inscription.fk_estudiante.apellido_1}, {obj.fk_inscription.fk_estudiante.nombre_1}'
    ws['B11'] = 'La Suma de:'
    ws['D11'] = f"{num2words(obj.monto_pagar, lang='es')} balboas."
    ws['B13'] = 'Facturacion Por:'
    ws['D13'] = f'{obj.cargo_tipo}'
    ws['B15'] = 'Saldo Anterior:'
    ws['E15'] = f'{obj.monto_valor}'
    ws['B17'] = 'Abono'
    ws['D17'] = f'{obj.monto_pagar}'
    ws['B19'] = 'Saldo Actual:'
    if (obj.monto_pagar-obj.monto_valor) > 0:
        ws['D19'] = f'{obj.monto_pagar-obj.monto_valor}'
    else:
        ws['E19'] = f'{obj.monto_pagar-obj.monto_valor}'
    ws['E21'] = 'Firma:'

    # Configuramos la respuesta HTTP con el archivo Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=recibo-pago-{obj.factura_pago}.xlsx'
    wb.save(response)

    return response


def print_pays(request, num_fact_id):
    obj = get_object_or_404(Pago, pk=num_fact_id)
    wb = Workbook()
    ws = wb.active

    # Escribimos los títulos de las columnas
    ws['B2'] = 'INSTITUTO GIORDANO BRUNO'
    ws.merge_cells('B2:F2')
    ws['B3'] = 'R.U.C. 443964-1-430566 D.V. 65'
    ws.merge_cells('B3:F3')
    ws['B4'] = 'E-Mail: institutogiordanobruno2004@gmail.com'
    ws.merge_cells('B4:F4')

    # seccion con variables
    ws['A7'] = f'{obj.fk_inscription.fk_estudiante.apellido_1}, {obj.fk_inscription.fk_estudiante.nombre_1}'
    ws['F7'] = f'{obj.fk_inscription.fk_estudiante.dni}'
    # FILTER_orm

    # FILTER_SQL
    consulta = f"select p.fecha_pago, i.trimestre,  ROUND(p.monto_valor, 2), ROUND(p.monto_pagar, 2), ROUND((p.monto_valor-p.monto_pagar), 2) as saldo from app_pago as p inner join app_inscription as i on fk_inscription_id = num_comprobante and num_comprobante = '{obj.fk_inscription.num_comprobante}' order by factura_pago desc;"
    # END FILTER_SQL

    # SQL_CONNECTION
    sql_connection = connection.cursor()
    sql_connection.execute(consulta)
    result = sql_connection.fetchall()
    sql_connection.close()

    # recorre la variable result eh introduce cada uno de sus item en celdas
    row_num = 8
    for row in result:
        ws.cell(row=row_num, column=1, value=row[0])
        ws.cell(row=row_num, column=2, value=f'{row[1]}')
        ws.cell(row=row_num, column=3, value=row[2])
        ws.cell(row=row_num, column=4, value=row[3])
        ws.cell(row=row_num, column=5, value=row[4])
        row_num += 1

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=pagos-inscripcion-{obj.fk_inscription.num_comprobante}.xlsx'
    wb.save(response)

    return response